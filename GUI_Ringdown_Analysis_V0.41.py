'''
************************************************************************
Identification of Electromechanical Modes Driven by Ringdown Analysis
Prony, Eigensystem Realization Algorithm, Matrix Pencil
Graphical User Interface

Jose Antonio de la O Serna
Mario Roberto Arrieta Paternina
Rodrigo David Reyes de Luna

June 2021
************************************************************************
'''

import sys, os
import csv
import xlrd
import math
import time
import openpyxl
import matplotlib

import numpy 				as np
import tkinter 				as tk
import matplotlib.pyplot 	as plt
import tkinter.scrolledtext as st

from datetime 							import datetime
from scipy.linalg 						import hankel
from openpyxl.styles 					import Alignment, Font
from tkinter 							import filedialog, messagebox, ttk
from scipy.fftpack 						import fft, fftfreq, fftshift
from matplotlib.backends.backend_tkagg 	import FigureCanvasTkAgg, NavigationToolbar2Tk




class Window1():

	# Main window

	global font_text
	font_text = 'Calibri'

	global color0, color1, color2, color3
	color0 = 'SteelBlue4'
	color1 = 'black'
	color2 = 'grey24'
	color3 = 'snow'


	def __init__(self, master):
		self.master = master
		self.master.title('Ringdown Analysis Toolbox')
		self.master.geometry('1300x800')
		self.master.configure(bg = color1)
		self.frame = tk.Frame(self.master)
		self.frame.pack()
		self.Create_Widgets()


	def Create_Widgets(self):

		# MASTER
		title = 'Identification of Electromechanical Modes Driven by Ringdown Analysis'
		self.lbl0 = tk.Label(self.master, text = title, font = (font_text, 14), fg = 'white', bg = color0, anchor = tk.CENTER)
		self.lbl0.pack(fill = 'x')

		self.lbl1 = tk.Label(self.master, text = 'i', font = ('Times', 14, 'italic'), fg = 'white', bg = color1, anchor = tk.CENTER)
		self.lbl1.place(relx = 1.00, rely = 0.0, height = 30, width = 30,  anchor = tk.NE)
		self.lbl1.bind('<Button-1>', self.Info)

		self.btn1 = tk.Button(self.master, text = 'File', font = (font_text, 10), command = self.click_btn1, anchor = tk.CENTER)
		self.btn1.place(relx = 0.012, y = 50.0, height = 24, relwidth = 0.070)

		self.ent1 = tk.Entry(self.master, fg = 'white', font = (font_text, 10), bg = color2)
		self.ent1.place(relx = 0.085, y = 50.0, height = 24, relwidth = 0.50)
		self.ent1.insert(0, 'Select data file...')

		# MASTER
		self.btn31 = tk.Button(self.master, text = 'CLEAR ALL', font = (font_text, 12), command = self.click_btn31, anchor = tk.CENTER)
		self.btn31.place(relx = 0.60, y = 50, height = 30, relwidth = 0.11)

		self.btn32 = tk.Button(self.master, text = 'RUN', font = (font_text, 12), command = self.click_btn32, anchor = tk.CENTER)
		self.btn32.place(relx = 0.72, y = 50, height = 30, relwidth = 0.11)

		self.btn33 = tk.Button(self.master, text = 'EXIT', font = (font_text, 12), command = self.click_btn33, anchor = tk.CENTER)
		self.btn33.place(relx = 0.84, y = 50, height = 30, relwidth = 0.11)

		# LABELFORM 2 - SETTINGS
		self.lbfrm2 = tk.LabelFrame(self.master, text = 'Settings', font = (font_text, 10), fg = 'white', bg = color1)
		self.lbfrm2.place(x = 10, y = 80, height = 82, relwidth = 0.980, anchor = tk.NW)

		self.btn2 = tk.Button(self.lbfrm2, text = 'Edit', font = (font_text, 10), command = self.click_btn2, anchor = tk.CENTER)
		self.btn2.place(relx = 0.004, y = 2, height = 24, relwidth = 0.070)

		self.val_W1ent2 = tk.StringVar()
		self.val_W1ent2.set('Edit signals...')
		self.ent2 = tk.Entry(self.lbfrm2, textvariable = self.val_W1ent2, fg = 'white', font = (font_text, 10), bg = color2)
		self.ent2.place(relx = 0.078, y = 2, height = 24, relwidth = 0.21)

		self.lbl5 = tk.Label(self.lbfrm2, text = 'Method:', font = (font_text, 10), fg = 'white', bg = color1,  anchor = tk.NE)
		self.lbl5.place(relx = 0.004, y = 32, height = 24, relwidth = 0.070, anchor = tk.NW)

		options2 = ['Prony', 'ERA', 'Matrix Pencil']
		self.val_opmn2 = tk.StringVar()
		self.val_opmn2.set(options2[0])
		self.opmn2 = tk.OptionMenu(self.lbfrm2, self.val_opmn2, *options2, command = self.Label_Method)
		self.opmn2.place(relx = 0.078, y = 32, height = 24, relwidth = 0.10, anchor = tk.NW)

		options3 = ['Single-Channel', 'Multi-Channel']
		self.val_opmn3 = tk.StringVar()
		self.val_opmn3.set(options3[0])
		self.opmn3 = tk.OptionMenu(self.lbfrm2, self.val_opmn3, *options3)
		self.opmn3.place(relx = 0.19, y = 32, height = 24, relwidth = 0.10, anchor = tk.NW)

		self.lbl2 = tk.Label(self.lbfrm2, text = 'Time interval [start][end]:', font = (font_text, 10), fg = 'white', bg = color1,  anchor = tk.E)
		self.lbl2.place(relx = 0.470, y = 2, height = 24, relwidth = 0.12, anchor = tk.NE)

		self.val_W1ent3 = tk.DoubleVar()
		self.ent3 = tk.Entry(self.lbfrm2, textvariable = self.val_W1ent3, fg = 'white', font = (font_text, 10), bg = color2)
		self.ent3.place(relx = 0.473, y = 2, height = 24, relwidth = 0.05)

		self.val_W1ent4 = tk.DoubleVar()
		self.ent4 = tk.Entry(self.lbfrm2, textvariable = self.val_W1ent4, fg = 'white', font = (font_text, 10), bg = color2)
		self.ent4.place(relx = 0.530, y = 2, height = 24, relwidth = 0.05)

		self.lbl3 = tk.Label(self.lbfrm2, text = 'Frequency interval (Hz.) [min][max]:', font = (font_text, 10), fg = 'white', bg = color1,  anchor = tk.E)
		self.lbl3.place(relx = 0.470, y = 32, height = 24, relwidth = 0.16, anchor = tk.NE)

		self.ent5 = tk.Entry(self.lbfrm2, fg = 'white', font = (font_text, 10), bg = color2)
		self.ent5.place(relx = 0.473, y = 32, height = 24, relwidth = 0.05)
		self.ent5.insert(0, 0.0)

		self.ent6 = tk.Entry(self.lbfrm2, fg = 'white', font = (font_text, 10), bg = color2)
		self.ent6.place(relx = 0.530, y = 32, height = 24, relwidth = 0.05)
		self.ent6.insert(0, 1.0)

		self.val_W1chb1 = tk.BooleanVar()
		self.val_W1chb1.set(False)
		self.chb1 = tk.Checkbutton(self.lbfrm2, variable = self.val_W1chb1, text = 'Sliding window [length]:', 
			onvalue = True, offvalue = False, fg = 'white', bg = color1, anchor = tk.NE)
		self.chb1.config(selectcolor='#000000')
		self.chb1.place(relx = 0.72, y = 2, height = 24, relwidth = 0.125, anchor = tk.NE)

		self.ent8 = tk.Entry(self.lbfrm2, fg = 'white', font = (font_text, 10), bg = color2)
		self.ent8.place(relx = 0.723, y = 2, height = 24, relwidth = 0.050)

		self.val_W1chb2 = tk.BooleanVar()
		self.val_W1chb2.set(False)
		self.chb2 = tk.Checkbutton(self.lbfrm2, variable = self.val_W1chb2, text = 'Downsampling [factor]:', 
			onvalue = True, offvalue = False, fg = 'white', bg = color1, anchor = tk.NE)
		self.chb2.config(selectcolor='#000000')
		self.chb2.place(relx = 0.72, y = 32, height = 24, relwidth = 0.125, anchor = tk.NE)

		self.val_W1ent9 = tk.IntVar()
		self.val_W1ent9.set(1)
		self.ent9 = tk.Entry(self.lbfrm2, textvariable = self.val_W1ent9, fg = 'white', font = (font_text, 10), bg = color2)
		self.ent9.place(relx = 0.723, y = 32, height = 24, relwidth = 0.050)

		self.lbl4 = tk.Label(self.lbfrm2, text = 'Modes:', font = (font_text, 10), fg = 'white', bg = color1,  anchor = tk.E)
		self.lbl4.place(relx = 0.893, y = 2, height = 24, relwidth = 0.11, anchor = tk.NE)

		self.val_W1ent7 = tk.DoubleVar()
		self.ent7 = tk.Entry(self.lbfrm2, textvariable = self.val_W1ent7, fg = 'white', font = (font_text, 10), bg = color2)
		self.ent7.place(relx = 0.895, y = 2, height = 24, relwidth = 0.050)

		self.btn4 = tk.Button(self.lbfrm2, text = 'View', font = (font_text, 10), command = self.click_btn4, anchor = tk.CENTER)
		self.btn4.place(relx = 0.95, y = 2, height = 24, relwidth = 0.045)
		self.Label_Method()

		self.val_W1chb3 = tk.BooleanVar()
		self.val_W1chb3.set(False)
		self.chb3 = tk.Checkbutton(self.lbfrm2, variable = self.val_W1chb3, text = 'Normalize signal', 
			onvalue = True, offvalue = False, fg = 'white', bg = color1, anchor = tk.NE)
		self.chb3.config(selectcolor='#000000')
		self.chb3.place(relx = 0.895, y = 32, height = 24, relwidth = 0.1, anchor = tk.NE)

		# LABELFORM 3 - RESULTS -------------------------------------------------------------------------------------------------------------------------
		self.lbfrm3 = tk.LabelFrame(self.master, text = 'Results', font = (font_text, 10), fg = 'white', bg = color1)
		self.lbfrm3.place(x = 10, y = 165, relheight = 0.77, relwidth = 0.980, anchor = tk.NW)

		self.lbl6 = tk.Label(self.lbfrm3, text = 'Last RUN (date-time): ...', font = (font_text, 10), fg = 'white', bg = color1,  anchor = tk.W)
		self.lbl6.place(relx = 0.00, rely = 0.00, relwidth = 0.25, anchor = tk.NW)

		self.lbl7 = tk.Label(self.lbfrm3, text = 'Method execution time (s): ', font = (font_text, 10), fg = 'white', bg = color1,  anchor = tk.W)
		self.lbl7.place(relx = 0.00, rely = 0.030, relwidth = 0.25, anchor = tk.NW)

		self.lbl8 = tk.Label(self.lbfrm3, text = 'Total execution time (s): ', font = (font_text, 10), fg = 'white', bg = color1,  anchor = tk.W)
		self.lbl8.place(relx = 0.00, rely = 0.060, relwidth = 0.25, anchor = tk.NW)

		self.btn5 = tk.Button(self.lbfrm3, text = 'Select All', font = (font_text, 10), command = lambda: self.click_btn56(True), anchor = tk.CENTER)
		self.btn5.place(relx = 0.002, rely = 0.12, height = 24, relwidth = 0.092, anchor = tk.NW)

		self.btn6 = tk.Button(self.lbfrm3, text = 'Deselect All', font = (font_text, 10), command = lambda: self.click_btn56(False), anchor = tk.CENTER)
		self.btn6.place(relx = 0.1, rely = 0.12, height = 24,  relwidth = 0.092, anchor = tk.NW)

		self.val_opmn1 = tk.StringVar()
		self.opmn1 = tk.OptionMenu(self.lbfrm3, self.val_opmn1, ())
		self.opmn1.place(relx = 0.198, rely = 0.12, height = 24,  relwidth = 0.092, anchor = tk.NW)

		# TREE VIEW FOR SHOW MODAL DATA
		trv1_columns 	= ('#1', '#2', '#3', '#4', '#5', '#6', '#7', '#8', '#9', '#10')
		trv1_name		= ('Selection',	'Mode', 'Type', 'Frequency', 'Amplitude', 'Damping', 'Damping Ratio', 'Phase', 'Pole-Zero', 'Energy')
		self.trv1 = ttk.Treeview(self.lbfrm3,  show='headings', selectmode = 'browse', columns = trv1_columns)
		self.trv1.place(relx = .002, rely = 0.175, relheight = 0.76, relwidth = .29, anchor = tk.NW)
		self.trv1.bind('<<TreeviewSelect>>', self.Selection_trv1)

		for index, item in enumerate(trv1_columns):
			self.trv1.column(item, width = 100, minwidth=50, anchor=tk.CENTER)
			self.trv1.heading(item, text = trv1_name[index], anchor=tk.CENTER)

		self.treescrollx = tk.Scrollbar(self.trv1, orient = 'horizontal', command = self.trv1.xview)
		self.treescrolly = tk.Scrollbar(self.trv1, orient = 'vertical', command = self.trv1.yview)
		self.trv1.configure(xscrollcommand = self.treescrollx.set, yscrollcommand = self.treescrolly.set)
		self.treescrolly.pack(side = 'right', fill = tk.Y)
		self.treescrollx.pack(side = 'bottom', fill = tk.X)

		self.btn7 = tk.Button(self.lbfrm3, text = 'Export', font = (font_text, 10), command = self.click_btn7, anchor = tk.CENTER)
		self.btn7.place(relx = 0.002, rely = 0.95, height = 24,  relwidth = 0.092, anchor = tk.NW)

		# NOTEBOOK FOR GRAPHICS
		self.tbctrl = ttk.Notebook(self.lbfrm3)
		self.tbctrl.place(relx = 0.297, rely = 0.00, relwidth = 0.70, relheight = 0.99, anchor = tk.NW)
		self.tab1 = tk.Frame(self.tbctrl, background = 'black')
		self.tab2 = tk.Frame(self.tbctrl, background = 'black')
		self.tab3 = tk.Frame(self.tbctrl, background = 'black')
		self.tab4 = tk.Frame(self.tbctrl, background = 'black')
		self.tbctrl.add(self.tab1, text ='Signals')
		self.tbctrl.add(self.tab2, text ='FFT')
		self.tbctrl.add(self.tab3, text ='Mode Shapes')
		self.tbctrl.add(self.tab4, text ='Pole-Zero')


	def Info(self, event = None):

		info_text =  ('''
Reference:
J. Sanchez-Gasca and D. Trudnowski
“Identification of electromechanicalmodes in power system”
IEEE Task Force on Identification of Electromechanical Modes of the Power System Stability, Power & Energy Society, Tech. Rep.
June 2012

Authors:
José Antonio de la O Serna
Mario Roberto Arrieta Paternina
Rodrigo David Reyes de Luna

Version 0.0
August 11th, 2021
		''')
		tk.messagebox.showinfo('Information', info_text)


	def click_btn1(self):

		# -------------------
		# |      FILE       |
		# -------------------

		self.ent1.delete(0, 'end')
		self.ent2.delete(0, 'end')
		self.ent3.delete(0, 'end')
		self.ent4.delete(0, 'end')

		file_path = filedialog.askopenfilename(title		='Select a file',
		                          	   		   filetype		=[('csv files', '*.csv'),('xlsx files', '*.xlsx')])
		self.ent1.insert(0, file_path)

		if self.Check_Data() == True:

			Data, h_vec, t_vec, y_vec =  self.Get_Data(file_path)

			h_vec = str(h_vec[1:]).replace("'", '')[1:-1]
			self.ent2.insert(0, h_vec)

			self.ent3.insert(0, t_vec[0])
			self.ent4.insert(0, t_vec[-1])


	def click_btn2(self):
		
		# -------------------
		# |      EDIT       |
		# -------------------

		if self.Check_Data() == True:

			file_path = self.ent1.get()

			Window2(tk.Toplevel(self.master), self.val_opmn3, self.val_W1ent2, self.val_W1ent3, self.val_W1ent4, 
											  self.val_W1chb2, self.val_W1ent9, self.val_W1chb3, file_path)


	def click_btn31(self):

		# -------------------
		# |    CLEAR ALL    |
		# -------------------

		self.ent1.delete(0, 'end')
		self.ent2.delete(0, 'end')
		self.ent3.delete(0, 'end')
		self.ent4.delete(0, 'end')
		self.ent5.delete(0, 'end')
		self.ent6.delete(0, 'end')
		self.ent7.delete(0, 'end')
		self.ent8.delete(0, 'end')
		self.ent9.delete(0, 'end')

		self.val_W1chb1.set(False)
		self.val_W1chb2.set(False)
		self.val_W1chb3.set(False)

		self.lbl6.config(text = 'Last RUN (date-time): ')
		self.lbl7.config(text = 'Method execution time (s): ')
		self.lbl8.config(text = 'Total execution time (s): ')

		self.val_opmn1 = tk.StringVar()
		self.opmn1 = tk.OptionMenu(self.lbfrm3, self.val_opmn1, ())
		self.opmn1.place(relx = 0.198, rely = 0.12, height = 24,  relwidth = 0.092, anchor = tk.NW)

		self.trv1.delete(*self.trv1.get_children())

		try:
			self.canvas_sig.get_tk_widget().destroy()
			self.toolbar_sig.destroy()
		except:
			pass

		try:
			self.canvas_fft.get_tk_widget().destroy()
			self.toolbar_fft.destroy()
		except:
			pass

		try:
			self.canvas_msh.get_tk_widget().destroy()
			self.toolbar_msh.destroy()
		except:
			pass

		try:
			self.canvas_rts.get_tk_widget().destroy()
			self.toolbar_rts.destroy()
		except:
			pass


	def click_btn32(self):

		# -------------------
		# |       RUN       |
		# -------------------

		if self.Check_Data(2, 34, 56, 7, 9) == True:

			t0_run = time.time()

			global M_res, M_rts, modes, L_mod, list_signals, t_vec, y_vec, dt, N, t_aprx, med, can

			list_signals	= self.List_Data_byComas(self.ent2.get())
			file_path 		= self.ent1.get()
			t_start 		= float(self.ent3.get())
			t_end 			= float(self.ent4.get())
			fct_method 		= float(self.ent7.get())

			self.opmn1.destroy()
			self.opmn1 = tk.OptionMenu(self.lbfrm3, self.val_opmn1, *list_signals, command = self.Load_trv1)
			self.opmn1.place(relx = 0.198, rely = 0.12, height = 24,  relwidth = 0.092, anchor = tk.NW)
			self.val_opmn1.set(list_signals[0])

			Data, h_vec, t_vec, y_vec = self.Get_Data(file_path)

			for i in range(y_vec.shape[1]):
				y_vec[: , i] = y_vec[: , i] - np.mean(y_vec[: , i])

			for i in range(len(h_vec[1:]), 0, -1):
				if h_vec[i] not in list_signals:
					y_vec = np.delete(y_vec, (i - 1), axis = 1)

			if self.val_W1chb2.get() == True:
				fctr_m = int(self.ent9.get())
				t_vec, y_vec = self.Downsampling(fctr_m, t_vec, y_vec)

			t0_method = time.time()

			t_start 	= float(t_start)
			t_end 		= float(t_end)

			dt = float(t_vec[1] - t_vec[0])
			pa = int(round(t_start / dt, 0))
			pb = int(round(t_end / dt, 0)) + int(1)

			t_vec = t_vec[pa : pb]
			y_vec = y_vec[pa : pb , :]

			N 		= len(t_vec)
			t_aprx 	= np.linspace(0, t_end - t_start, N)
			med 	= y_vec.shape[0]
			can 	= y_vec.shape[1]

			L_res = []
			L_mod = []
			
			if self.val_opmn3.get() == 'Single-Channel':

				if self.val_W1chb3.get() == True:
					y_vec = self.Normalize(y_vec, 1)

				for c in range(can):

					Y = np.zeros([y_vec.shape[0], 1])
					Y = (y_vec.T)[c]

					if self.val_opmn2.get() == 'Prony':
						modes, mag, ang, damp, freq, damprat, enrgy, roots = self.Prony(Y, N, dt, fct_method, t_aprx, t_vec)

					if self.val_opmn2.get() == 'Matrix Pencil':
						modes, mag, ang, damp, freq, damprat, enrgy, roots = self.MatrixPencil(Y, N, dt, fct_method, t_aprx, t_vec)

					if self.val_opmn2.get() == 'ERA':
						modes, mag, ang, damp, freq, damprat, enrgy, roots = self.ERA(Y, N, dt, fct_method, t_aprx, t_vec)

					L_res.append([mag, ang, damp, freq, damprat, enrgy, roots])
					L_mod.append(modes)


			if self.val_opmn3.get() == 'Multi-Channel':

				if self.val_W1chb3.get() == True:
					y_vec = self.Normalize(y_vec, 2)

				Y = y_vec
				if self.val_opmn2.get() == 'ERA':
					modes, mag, ang, damp, freq, damprat, enrgy, roots = self.MultiERA(Y, N, dt, fct_method, t_aprx, t_vec)

					L_res.append([mag, ang, damp, freq, damprat, enrgy, roots])
					L_mod.append(modes)

			t1_method = time.time()

			max_mod = max(L_mod)
			M_res 	= np.zeros([max_mod, can, 6])
			M_rts 	= np.zeros([max_mod, can], dtype = complex)

			if self.val_opmn3.get() == 'Single-Channel':
				for c in range(can):
					for m in range(L_mod[c]):
						M_res[m , c , 0] 	= L_res[c][0][m]	# mag
						M_res[m , c , 1] 	= L_res[c][1][m]	# ang
						M_res[m , c , 2] 	= L_res[c][2][m]	# damp
						M_res[m , c , 3] 	= L_res[c][3][m]	# freq
						M_res[m , c , 4] 	= L_res[c][4][m]	# damprat
						M_res[m , c , 5] 	= L_res[c][5][m]	# enrgy
						M_rts[m , c] 		= L_res[c][6][m]	# roots


			if self.val_opmn3.get() == 'Multi-Channel':
				M_res[: , : , 0] 	= mag
				M_res[: , : , 1] 	= ang
				M_res[: , : , 2] 	= damp
				M_res[: , : , 3] 	= freq
				M_res[: , : , 4] 	= damprat
				M_res[: , : , 5] 	= enrgy
				M_rts[: , :] 		= roots


			self.Load_trv1()
			self.Fig_FFT()

			t1_run = time.time()

			self.lbl6.config(text = 'Last RUN (date-time): ' + datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
			self.lbl7.config(text = 'Method execution time (s): ' + str(t1_method - t0_method))
			self.lbl8.config(text = 'Total execution time (s): ' + str(t1_run - t0_run))


	def click_btn33(self):

		# -------------------
		# |      EXIT       |
		# -------------------

		self.master.destroy()


	def click_btn4(self):
		
		# -------------------
		# |      VIEW       |
		# -------------------

		if self.Check_Data(2, 34, 7, 9) == True:

			file_path 		= self.ent1.get()
			list_signals	= self.List_Data_byComas(self.ent2.get())
			i_channels 		= self.val_opmn3.get()
			t_start 		= self.ent3.get()
			t_end 			= self.ent4.get()
			b_dwnspl		= self.val_W1chb2.get()
			if b_dwnspl == True 	: f_dwnspl = int(self.ent9.get())
			if b_dwnspl == False 	: f_dwnspl = 0
			energy_th 		= float(self.val_W1ent7.get())

			Data, h_vec, t_vec, y_vec = self.Get_Data(file_path)

			Window3(tk.Toplevel(self.master), self.val_W1ent7, list_signals, i_channels, t_start, t_end, b_dwnspl, f_dwnspl, energy_th, h_vec, t_vec, y_vec)


	def click_btn56(self, value):

		# -----------------------
		# |    DE/SELECT ALL    |
		# -----------------------

		if len(self.trv1.get_children()) == 0: return

		for index, item in enumerate(self.trv1.get_children()):

			row_data = self.trv1.item(item)['values']
			row_text = self.trv1.item(item)['text']

			if row_text == (not value):

				if value == True	: row_data[0] = '✓'
				if value == False	: row_data[0] = ''

				self.trv1.delete(item)
				self.trv1.insert('', index, text = value, values = row_data)

		self.Load_Figures()


	def click_btn7(self):

		# --------------------
		# |      EXPORT      |
		# --------------------
		
		now = datetime.now()
		book_name = 'results_' + str(now.strftime('%d%m%Y_%H%M%S'))

		wb = openpyxl.Workbook()


		sheet_1 = wb['Sheet']
		sheet_1.title = 'Sheet1'

		sheet_1.column_dimensions['A'].width = 20

		sheet_1.column_dimensions['D'].width = 20
		sheet_1.column_dimensions['E'].width = 20
		sheet_1.column_dimensions['F'].width = 20
		sheet_1.column_dimensions['G'].width = 20
		sheet_1.column_dimensions['H'].width = 20
		sheet_1.column_dimensions['I'].width = 20
		sheet_1.column_dimensions['J'].width = 20
		sheet_1.column_dimensions['K'].width = 20

		# SETTINGS .....................................................................................

		sheet_1['A1'] 	= 'Identification of Electromechanical Modes Driven by Ringdown Analysis'
		sheet_1['A1'].font = Font(size = 16, bold = True)

		sheet_1['A2'] 	= 'Time-Date: '
		sheet_1['A2'].alignment = Alignment(horizontal = 'right')

		sheet_1['A6'] 	= 'Settings'
		sheet_1['A6'].font = Font(color = '00FF0000', bold = True)

		sheet_1['A7'] 	= 'File name: '
		sheet_1['A8'] 	= 'Signals: '
		sheet_1['A9'] 	= 'Method: '
		sheet_1['A10'] 	= 'Time interval: '
		sheet_1['A11'] 	= 'Frequency interval: '
		sheet_1['A12'] 	= 'Sliding window: '
		sheet_1['A13'] 	= 'Downsampling: '

		if self.val_opmn2.get() == 'Prony':
			sheet_1['A14'] 	= 'Modes: '
		else:
			sheet_1['A14'] 	= 'Energy threshold: '

		sheet_1['A15'] 	= 'Normalize signal: '

		sheet_1['B2'] 	= str(now.strftime('%d%b%Y - %H:%M:%S'))
		sheet_1['B7'] 	= str(self.ent1.get())
		sheet_1['B8'] 	= str(self.ent2.get())
		sheet_1['B9'] 	= self.val_opmn2.get() + ' ' + self.val_opmn3.get()
		sheet_1['B10'] 	= 'start: ' + str(self.ent3.get()) + ' / end: ' + str(self.ent4.get())
		sheet_1['B11'] 	= 'min: ' + str(self.ent5.get()) + ' / max: ' + str(self.ent6.get())

		if self.val_W1chb1.get() == True:
			sheet_1['B12'] 	= 'True - ' + str(self.ent8.get())
		else:
			sheet_1['B12'] 	= 'False - ' + str(self.ent8.get())

		if self.val_W1chb2.get() == True:
			sheet_1['B13'] 	= 'True - ' + str(self.ent9.get())
		else:
			sheet_1['B13'] 	= 'False - ' + str(self.ent9.get())

		if self.val_opmn2.get() == 'Prony':
			sheet_1['B14'] 	= str(self.ent7.get())
		else:
			sheet_1['B14'] 	= str(self.ent7.get()) + '%'
		
		if self.val_W1chb3.get() == True:
			sheet_1['B15'] 	= 'True'
		else:
			sheet_1['B15'] 	= 'False'

		for i in range(9):
			sheet_1.cell(7 + i , 1).alignment = Alignment(horizontal = 'right')

		# RESULTS .....................................................................................

		sheet_1['A18'] 	= 'Results'
		sheet_1['A18'].font = Font(color = '00FF0000', bold = True)

		sheet_1['A19'] 	= 'Method time (s): '
		sheet_1['A19'].alignment = Alignment(horizontal = 'right')

		sheet_1['A20'] 	= 'Total time (s): '
		sheet_1['A20'].alignment = Alignment(horizontal = 'right')

		sheet_1['B19'] 	= self.lbl7['text'][27 :]
		sheet_1['B20'] 	= self.lbl8['text'][26 :]

		list_title = ['Signal', 'Mode', 'Type', 'Frequency', 'Amplitude', 'Damping', 'Damping ratio', 'Phase', 'Pole-Zero', 'Energy']
		for index, title in enumerate(list_title):
			sheet_1.cell(22 , 2 + index).value 	= title
			sheet_1.cell(22 , 2 + index).font 	= Font(bold = True)


		jj = 0
		for i_signal in range(M_res.shape[1]):

			mag 	= M_res[: , i_signal , 0]
			ang 	= M_res[: , i_signal , 1]
			damp 	= M_res[: , i_signal , 2]
			freq 	= M_res[: , i_signal , 3]
			damprat	= M_res[: , i_signal , 4]
			enrgy 	= M_res[: , i_signal , 5]
			roots 	= M_rts[: , i_signal]

			m = 0
			for i in range(M_res.shape[0]):
				if freq[i] > 0:
					m = m + 1
					j = m - 1
					
					osc_mod = '-'
					if freq[i] >= 0.10 and freq[i] < 0.80: osc_mod = 'Inter-Area'
					if freq[i] >= 0.80 and freq[i] < 2.00: osc_mod = 'Local'
					if freq[i] >= 2.00 and freq[i] < 3.00: osc_mod = 'Intra-Plant'

					sheet_1.cell(23 + j + jj , 2).value = list_signals[i_signal]
					sheet_1.cell(23 + j + jj , 3).value = m
					sheet_1.cell(23 + j + jj , 4).value = osc_mod
					sheet_1.cell(23 + j + jj , 5).value = freq[i]
					sheet_1.cell(23 + j + jj , 6).value = mag[i]
					sheet_1.cell(23 + j + jj , 7).value = damp[i]
					sheet_1.cell(23 + j + jj , 8).value = damprat[i]
					sheet_1.cell(23 + j + jj , 9).value = ang[i]
					sheet_1.cell(23 + j + jj , 10).value = str(roots[i].real) +', '+ str(roots[i].imag)
					sheet_1.cell(23 + j + jj , 11).value = enrgy[i]

			jj = jj + j + 1


		for i in range(10):
			for j in range(jj + 1):
				sheet_1.cell(22 + j , 2 + i).alignment = Alignment(horizontal = 'center')

		wb.save(book_name + '.xlsx')


	def Load_Figures(self, *args):

		self.Fig_Signals()
		self.Fig_Roots()

		if self.val_opmn3.get() == 'Multi-Channel':
			self.Fig_ModalShape()


	def Fig_Roots(self):

		for index, i in enumerate(list_signals):
			if i == self.val_opmn1.get():
				i_signal = index
				break

		freq 	= M_res[: , i_signal , 3]
		roots 	= M_rts[: , i_signal]

		dic_item = {}
		for index, item in enumerate(self.trv1.get_children()):
			dic_item[index] = item

		dic_mode = {}
		m = 0
		for i in range(max(L_mod)):
			if freq[i] > 0:
				dic_mode[i] = m
				m = m + 1

		r_aprx = []
		for i in range(max(L_mod)):
			if (freq[i] > 0) and (self.trv1.item(dic_item[dic_mode[i]])['text'] == True):
				r_aprx.append([roots[i], dic_mode[i] + 1])
		
		plt.style.use('dark_background')
		self.fig = plt.Figure(dpi = 100) 
		ax = self.fig.add_subplot(111)

		ax.axvline(0, color = 'w', linestyle = 'dashed', lw = '1')
		ax.axhline(0, color = 'w', linestyle = 'dashed', lw = '1')

		self.unit_circle = plt.Circle((0, 0), 1.0, linestyle='dashed', color='w', fill=False)
		ax.add_patch(self.unit_circle)

		colors 	= plt.cm.get_cmap('gist_rainbow', can+1)
		for r in r_aprx:

			if abs(complex(r[0].real, r[0].imag)) > 1.0:
				mrk = 'x'
			else:
				mrk = 'o'

			ax.plot(r[0].real, +r[0].imag, linestyle = 'None', color = colors(i_signal), marker = mrk)
			ax.plot(r[0].real, -r[0].imag, linestyle = 'None', color = colors(i_signal), marker = mrk)
			ax.annotate(' ' + str(r[1]), (r[0].real, +r[0].imag), fontsize = 7)
			ax.annotate(' ' + str(r[1]), (r[0].real, -r[0].imag), fontsize = 7)

		ax.axis('equal')
		ax.set_title('Pole-Zero', fontsize = 12)
		ax.set_xlabel('Re')
		ax.set_ylabel('Im')
		ax.grid(True, color = 'grey',  linestyle = ':', linewidth = 0.75)

		self.fig.subplots_adjust(left = 0.08, bottom = 0.16, right = 0.96, top = 0.93)

		try:
			self.canvas_rts.get_tk_widget().destroy()
			self.toolbar_rts.destroy()
		except:
			pass

		self.canvas_rts = FigureCanvasTkAgg(self.fig, self.tab4)
		self.canvas_rts.draw() 
		self.canvas_rts.get_tk_widget().place(relheight = 1.00, relwidth = 1.00)
		self.toolbar_rts = NavigationToolbar2Tk(self.canvas_rts, self.tab4)
		self.toolbar_rts.place(relx = 0.50, rely = 1.00, relwidth = 1.00, anchor = tk.S)
		self.toolbar_rts.update()


	def Fig_ModalShape(self):

		try:
			self.canvas_msh.get_tk_widget().destroy()
			self.toolbar_msh.destroy()
		except:
			pass

		n = 0
		dic_trv1 = {}
		for index, item in enumerate(self.trv1.get_children()):
			row_text = self.trv1.item(item)['text']
			if row_text == True: n = n + 1
			dic_trv1[index] = {'item': item, 'text': row_text}

		if n == 0: return

		freq = M_res[: , 0 , 3]
		dic_mode = {}
		m = 0
		for i in range(len(freq)):
			if freq[i] > 0:
				dic_mode[i] = m
				m = m + 1

		#------------------------------------------------------------
		if n >=1: ax_pos = [[1 , 1]]
		if n >=2: ax_pos = [[1 , 2]]
		if n >=3: ax_pos = [[1 , 3]]

		ax_sw 	= True

		for index, i in enumerate(range(n), start = 1):

			ax_pos.append([ax_pos[i][0], ax_pos[i][1]])

			ax_sw = not ax_sw
			if ax_sw == False	: ax_pos[index][0] = ax_pos[i][0] + 1
			if ax_sw == True	: ax_pos[index][1] = ax_pos[i][1] + 1
		
		w = 0
		while n > ax_pos[w][0] * ax_pos[w][1]:
			w = w + 1
		#------------------------------------------------------------

		plt.style.use('dark_background')
		self.fig = plt.Figure(dpi = 100) 
		colors = plt.cm.get_cmap('hsv', can+1)

		im = 1
		for index, ifreq in enumerate(freq):

			if (freq[index] > 0) and (dic_trv1[dic_mode[index]]['text'] == True):

				imag	= M_res[index , : , 0]
				iang	= M_res[index , : , 1]

				imag_max 	= max(imag)
				index_imag	= np.where(imag == imag_max)
				iang_max 	= iang[index_imag[0][0]]

				ax = self.fig.add_subplot(ax_pos[w][0] , ax_pos[w][1] , im, projection = 'polar')
				for c in range(can):

					m = imag[c] / imag_max
					a = iang[c] - iang_max
					ax.quiver(a * (180.0 / np.pi), 0, 0, m, label = list_signals[c], color = colors(c), angles = 'xy', scale_units = 'xy', scale = 1.0)

				ax.set_rlim(0, 1.0)
				ax.set_rticks([0.25, 0.5, 0.75, 1.00])
				ax.set_yticklabels([])
				ax.set_rlabel_position(-90.0)
				ax.set_xticks(np.arange(0, 2.0 * np.pi, np.pi / 4.0))
				ax.set_xlabel('Mode ' + str(dic_mode[index] + 1) + ' / Freq ' + str(round(freq[index],3)))
				ax.grid(True, color = 'grey',  linestyle = ':', linewidth = 0.75)

				im = im + 1

		handles, labels = ax.get_legend_handles_labels()
		self.fig.legend(handles, labels, loc = 'lower center', bbox_to_anchor = (0.5, 0.05), fancybox = True, ncol = can)
		self.fig.subplots_adjust(left = 0.08, bottom = 0.16, right = 0.96, top = 0.93, wspace = 0.35, hspace = 0.35)

		self.canvas_msh = FigureCanvasTkAgg(self.fig, self.tab3)
		self.canvas_msh.draw() 
		self.canvas_msh.get_tk_widget().place(relheight = 1.00, relwidth = 1.00)
		self.toolbar_msh = NavigationToolbar2Tk(self.canvas_msh, self.tab3)
		self.toolbar_msh.place(relx = 0.50, rely = 1.00, relwidth = 1.00, anchor = tk.S)
		self.toolbar_msh.update()


	def Prony(self, Y, N, dt, fct_method, t_aprx, t_vec):

		modes = int(fct_method)
		if (modes == 0) or (modes > int(Y.shape[0] / 2)):
			modes = int(Y.shape[0] / 2)
			self.val_W1ent7.set(modes)

		T = np.zeros((N - modes, modes))
		for i in range(modes):
			T[:, i] = Y[modes-1-i : N-1-i]

		b = Y[modes : N]
		a = np.linalg.inv(T.T @ T) @ T.T @ b

		z0 = [1]
		for i in a:
			z0.append(-i)

		z = np.roots(z0)
		Z = np.zeros((N, modes), dtype = complex)

		for i in range(N):
			Z[i, :] = pow(z, i)

		lam = np.log(z)/dt

		B = np.linalg.inv(Z.T @ Z) @ Z.T @ Y

		mag 	= 2.0 * abs(B)
		ang 	= np.angle(B)
		damp 	= lam.real
		freq 	= lam.imag / (2.0 * np.pi)

		freq_zrs = list(np.where(freq == 0.0)[0])
		for fz in freq_zrs:
			freq[fz] = -0.0001

		omga	= 2.0 * np.pi / freq
		damprat	= (damp / omga) * 100.0
		enrgy 	= (1.0 / 2.0) * (omga**2)  * (mag**2)
		roots	= z

		return modes, mag, ang, damp, freq, damprat, enrgy, roots


	def MatrixPencil(self, Y, N, dt, fct_method, t_aprx, t_vec):

		r = int(np.around((N/2.0)-1.0, 0))
		Hankel = hankel(Y)
		H0 = Hankel[0:r, 0:r]
		H1 = Hankel[1:r+1, 0:r]

		u, s0, v0 = np.linalg.svd(H0, full_matrices=False)
		v = v0.T

		sum_s = sum(s0)
		sum_st = 0.0
		for modes, i in enumerate(s0, start = 1):
			sum_st 		= sum_st + i
			pc_sum_st 	= (sum_st / sum_s) * 100.0
			if pc_sum_st > float(fct_method): break

		V1 = np.zeros((r - 1 , modes))
		V2 = np.zeros((r - 1 , modes))

		for i in range(modes):
			V1[: , i] = v[0 : r - 1 , i]
			V2[: , i] = v[1 : r , i]

		Y1 = np.dot(V1.T , V1)
		Y2 = np.dot(V2.T , V1)

		z = np.linalg.inv(Y1)
		z = np.dot(z , Y2)
		z, b = np.linalg.eig(z)

		# MPM finished, continue as Prony method...
		lam = np.log(z) / dt

		Z = np.zeros((N, modes), dtype = complex)
		for i in range(N):
			Z[i, :] = pow(z, i)

		B = np.linalg.inv(Z.T @ Z) @ Z.T @ Y

		mag 	= 2.0 * abs(B)
		ang 	= np.angle(B)
		damp 	= lam.real
		freq 	= lam.imag / (2.0 * np.pi)

		freq_zrs = list(np.where(freq == 0.0)[0])
		for fz in freq_zrs:
			freq[fz] = -0.0001

		omga	= 2.0 * np.pi / freq
		damprat	= (damp / omga) * 100.0
		enrgy 	= (1.0 / 2.0) * (omga**2)  * (mag**2)
		roots	= z

		return modes, mag, ang, damp, freq, damprat, enrgy, roots


	def ERA(self, Y, N, dt, fct_method, t_aprx, t_vec):

		r = int(np.around((N/2.0)-1.0, 0))
		Hankel = hankel(Y)
		H0 = Hankel[0:r, 0:r]
		H1 = Hankel[1:r+1, 0:r]

		u, s0, v0 = np.linalg.svd(H0, full_matrices=False)
		s = np.diag(s0)
		v = v0.T

		sum_s 	= np.sum(s0)
		sum_st 	= 0.0
		for modes, i in enumerate(s0,  start = 1):
			sum_st 		= sum_st + i
			pc_sum_st 	= (sum_st / sum_s) * 100.0
			if pc_sum_st > float(fct_method): break

		U = u[: , 0:modes]
		S = s[0:modes , 0:modes]
		V = v[: , 0:modes]

		Sr = np.zeros((modes , modes))
		for i in range(modes):
			Sr[i,i] = pow(S[i,i], -0.5)

		A1 = Sr @ U.T @ H1 @ V @ Sr

		z, b = np.linalg.eig(A1)

		Z = np.zeros((N, modes), dtype = complex)
		for i in range(N):
			Z[i, :] = pow(z, i)

		# ERA finished, continue as Prony method...
		lam = np.log(z) / dt

		Z = np.zeros((N, modes), dtype = complex)
		for i in range(N):
			Z[i, :] = pow(z, i)

		B = np.linalg.inv(Z.T @ Z) @ Z.T @ Y

		mag 	= 2.0 * abs(B)
		ang 	= np.angle(B)
		damp 	= lam.real
		freq 	= lam.imag / (2.0 * np.pi)

		freq_zrs = list(np.where(freq == 0.0)[0])
		for fz in freq_zrs:
			freq[fz] = -0.0001

		omga	= 2.0 * np.pi / freq
		damprat	= (damp / omga) * 100.0
		enrgy 	= (1.0 / 2.0) * (omga**2)  * (mag**2)
		roots	= z

		return modes, mag, ang, damp, freq, damprat, enrgy, roots


	def MultiERA(self, Y, N, dt, fct_method, t_aprx, t_vec):

		r 	= int(np.around((med / 2.0) - 1.0, 0))

		H0 = np.zeros([r * can, r])
		H1 = np.zeros([r * can, r])
		for j in range(r):
			for i in range(r):
				H0[can*i: can*(i+1) , j] = Y[j+i+1 , :]
				H1[can*i: can*(i+1) , j] = Y[j+i+2 , :]

		u, s0, v0 = np.linalg.svd(H0, full_matrices=True)
		s = np.diag(s0)
		v = v0.T

		sum_s = sum(s0)
		sum_st = 0.0
		for modes, i in enumerate(s0, start = 1):
			sum_st 		= sum_st + i
			pc_sum_st 	= (sum_st / sum_s) * 100.0
			if pc_sum_st > float(fct_method): break
		
		U = u[: , 0:modes]
		S = s[0:modes , 0:modes]
		V = v[: , 0:modes]

		Sr = np.zeros((modes , modes))
		for i in range(modes):
			Sr[i,i] = pow(S[i,i], -0.5)

		A = Sr @ U.T @ H1 @ V @ Sr

		eigA, z = np.linalg.eig(A)
		lam = np.log(eigA)/dt

		Z = np.zeros((N, modes), dtype = complex)
		for i in range(N):
			Z[i, :] = pow(eigA, i)

		B = np.dot(np.linalg.pinv(Z), Y)

		# Results ------------------------------------------------->
		roots0 	= eigA
		damp0 	= lam.real
		freq0 	= lam.imag / (2.0 * np.pi)
		mag 	= 2 * abs(B)
		ang 	= np.angle(B)

		# Change any zero in frequency vector by -0.0001
		zrs = list(np.where(freq0 == 0.0)[0])
		for z in zrs: freq0[z] = -0.0001

		omga	= 2.0 * np.pi * freq0
		damprat0	= (damp0 / omga) * 100.0

		enrgy = np.zeros((len(freq0) , can))
		for i in range(len(freq0)):
			enrgy[i , :] = (1.0 / 2.0) * (omga[i]**2)  * (mag[i , :]**2)


		freq 	= np.zeros([modes , can])
		damp 	= np.zeros([modes , can])
		damprat = np.zeros([modes , can])
		roots 	= np.zeros([modes , can], dtype = complex)

		for i in range(can):
			damp[: , i] 	= damp0
			freq[: , i] 	= freq0
			damprat[: , i] 	= damprat0
			roots[: , i] 	= roots0
		
		return modes, mag, ang, damp, freq, damprat, enrgy, roots


	def Fig_FFT(self):

		y_z 	= np.zeros([2**14 , 1]) 
		medf 	= len(y_vec[:,0])

		S0 = np.zeros([len(y_z) + medf , can])
		Es = np.zeros([len(y_z) + medf , can], dtype = complex)

		for i in range(can):
			S0[: , i] = np.concatenate((y_vec[: , i], y_z), axis = None)
			Es[: , i] = fftshift(fft(S0[: , i]))

		fx = np.linspace(-1.0 / (2.0 * dt), 1.0 / (2.0 * dt), len(S0[: , 0]))

		try:
			self.canvas_fft.get_tk_widget().destroy()
			self.toolbar_fft.destroy()
		except:
			pass

		plt.style.use('dark_background')
		self.fig = plt.Figure(dpi = 100) 
		ax = self.fig.add_subplot(111)
		ax.set_xlabel('Frequency (Hz.)')
		ax.set_ylabel('Magnitude')
		ax.set_xlim(float(self.ent5.get()), float(self.ent6.get()))
		self.fig.subplots_adjust(left = 0.09, bottom = 0.16, right = 0.90, top = 0.95)
		self.canvas_fft = FigureCanvasTkAgg(self.fig, self.tab2)
		self.canvas_fft.draw() 
		self.canvas_fft.get_tk_widget().place(relheight = 1.00, relwidth = 1.00)
		self.toolbar_fft = NavigationToolbar2Tk(self.canvas_fft, self.tab2)
		self.toolbar_fft.place(relx = 0.50, rely = 1.00, relwidth = 1.00, anchor = tk.S)
		self.toolbar_fft.update()

		colors = plt.cm.get_cmap('gist_rainbow', can+1)
		for i in range(can):
			ax.plot(fx, abs(Es[: , i]) / (2**14), color = colors(i), label = list_signals[i], lw = '2')
		
		# ax.legend(loc = 'center right', bbox_to_anchor = (1.12, 0.5))
		ax.legend(loc = 'upper right')
		ax.grid(True, color = 'grey',  linestyle = ':', linewidth = 0.75)


	def Load_trv1(self, *args):

		for index, i in enumerate(list_signals):
			if i == self.val_opmn1.get():
				i_signal = index
				break

		self.trv1.delete(*self.trv1.get_children())

		mag 	= M_res[: , i_signal , 0]
		ang 	= M_res[: , i_signal , 1]
		damp 	= M_res[: , i_signal , 2]
		freq 	= M_res[: , i_signal , 3]
		damprat	= M_res[: , i_signal , 4]
		enrgy 	= M_res[: , i_signal , 5]
		roots 	= M_rts[: , i_signal]

		m = 1
		for i in range(max(L_mod)):
			if freq[i] > 0:
				
				osc_mod = '-'
				if freq[i] >= 0.10 and freq[i] < 0.80: osc_mod = 'Inter-Area'
				if freq[i] >= 0.80 and freq[i] < 2.00: osc_mod = 'Local'
				if freq[i] >= 2.00 and freq[i] < 3.00: osc_mod = 'Intra-Plant'

				row_val = ('✓', m, osc_mod, round(freq[i]	, 4),
											round(mag[i]	, 4),
											round(damp[i]	, 4),
											round(damprat[i], 1),
											round(ang[i]	, 4),
											complex(round(roots[i].real, 4), round(roots[i].imag, 4)),
											round(enrgy[i]	, 4))

				self.trv1.insert('', i, text = True, values = row_val)
				m = m + 1

		self.Load_Figures()


	def Selection_trv1(self, event):

		dic_item = {}
		for index, item in enumerate(self.trv1.get_children()):
			dic_item[index] = item

		index       = self.trv1.index(self.trv1.selection())
		row_data    = self.trv1.item(dic_item[index])['values']
		row_text    = self.trv1.item(dic_item[index])['text']

		if row_text == True:
			row_data[0] = ''
			text = False
		else:
			row_data[0] = '✓'
			text = True

		self.trv1.delete(dic_item[index])
		self.trv1.insert('', index, text = text, values = row_data)

		self.Load_Figures()


	def Fig_Signals(self):

		for index, i in enumerate(list_signals):
			if i == self.val_opmn1.get():
				i_signal = index
				break

		mag 	= M_res[: , i_signal , 0]
		ang 	= M_res[: , i_signal , 1]
		damp 	= M_res[: , i_signal , 2]
		freq 	= M_res[: , i_signal , 3]

		dic_item = {}
		for index, item in enumerate(self.trv1.get_children()):
			dic_item[index] = item

		dic_mode = {}
		m = 0
		for i in range(max(L_mod)):
			if freq[i] > 0:
				dic_mode[i] = m
				m = m + 1

		y_aprx = np.zeros((N))
		for i in range(max(L_mod)):
			if (freq[i] > 0) and (self.trv1.item(dic_item[dic_mode[i]])['text'] == True):
				y_aprx = y_aprx + mag[i] * np.exp(damp[i] * t_aprx) * np.cos((2 * np.pi * freq[i]) * t_aprx + ang[i])
		
		try:
			self.canvas_sig.get_tk_widget().destroy()
			self.toolbar_sig.destroy()
		except:
			pass

		plt.style.use('dark_background')
		plt.rcParams.update({'font.size': 9})
		self.fig = plt.Figure(dpi = 100) 
		ax = self.fig.add_subplot(111)
		ax.set_xlabel('Time (s)')
		ax.set_ylabel('Magnitude')
		ax.set_facecolor('k')
		self.fig.subplots_adjust(left = 0.08, bottom = 0.16, right = 0.96, top = 0.93)
		self.canvas_sig = FigureCanvasTkAgg(self.fig, self.tab1)
		self.canvas_sig.draw() 
		self.canvas_sig.get_tk_widget().place(relheight = 1.00, relwidth = 1.00)
		self.toolbar_sig = NavigationToolbar2Tk(self.canvas_sig, self.tab1)
		self.toolbar_sig.place(relx = 0.50, rely = 1.00, relwidth = 1.00, anchor = tk.S)
		self.toolbar_sig.update()

		colors 	= plt.cm.get_cmap('gist_rainbow', can+1)
		s0 		= np.linalg.norm(list(y_vec[: , i_signal]))
		s1 		= np.linalg.norm(list(y_aprx))
		sfull2 	= pow(s0 , 2)
		e2 		= pow(s0 - s1 , 2)
		SNR 	= round(10 * math.log10(sfull2 / e2) , 2)

		ax.plot(t_vec, y_vec[: , i_signal], color = colors(i_signal), label = list_signals[i_signal] + ': ' + str(SNR), lw = '2')
		ax.plot(t_vec, y_aprx, dashes = [6,4], color = colors(i_signal), lw = '2')
		ax.legend(loc = 'upper right', title = 'Signal Noise Ratio [dB]')
		ax.set_title('Method: ' + self.val_opmn2.get() + ' / ' + self.val_opmn3.get())
		ax.grid(True, color = 'grey',  linestyle = ':', linewidth = 0.75)


	def List_Data_byComas(self, str_signals):

		str_signals = str_signals.replace(' ', '')

		if len(str_signals) == 0: return False

		if str_signals[0] 	== ',': str_signals = str_signals[1:]
		if str_signals[-1] 	== ',': str_signals = str_signals[:-1]

		list_signals = []

		a = 0
		for i in range(len(str_signals)):
			if str_signals[i] == ',':
				list_signals.append(str_signals[a : i])
				a = i + 1

		list_signals.append(str_signals[a : len(str_signals)])

		return list_signals


	def Check_Data(self, *args):

		file_path 	= self.ent1.get()
		str_signals = self.ent2.get()
		t_start 	= self.ent3.get()
		t_end 		= self.ent4.get()
		f_min		= self.ent5.get()
		f_max		= self.ent6.get()
		fct_method	= self.ent7.get()
		dwn_spl		= self.ent9.get()


		if os.path.isfile(file_path) == False:
			tk.messagebox.showerror('Error', 'The file you have chosen is not valid.')
			return False

		if ((file_path.endswith('.csv') == False) and (file_path.endswith('.xlsx') == False)):
			tk.messagebox.showerror('Error', 'The file you have chosen is not valid. Must be extension csv or xlsx.')
			return False

		Data, h_vec, t_vec, y_vec =  self.Get_Data(file_path)

		if len(Data) == 0:
			tk.messagebox.showerror('Error', 'The file you have chosen is empty.')
			return False

		if len(y_vec) == 0:
			tk.messagebox.showerror('Error', 'Invalid format. See Details.')
			return False


		if 2 in args:
			
			list_signals = self.List_Data_byComas(str_signals)

			if list_signals == False:
				tk.messagebox.showerror('Error', 'Enter the name of the signals to be analyzed')
				return False

			for i in list_signals:
				if i not in h_vec:
					tk.messagebox.showerror('Error', 'The name signals are not valid. Check "' + str(i) + '"')
					return False

		if 34 in args:

			try:
				t_start = float(t_start)
				t_end 	= float(t_end)
			except:
				tk.messagebox.showerror('Error', 'Time interval data are not valid.')
				return False

			if (t_start < t_vec[0]) or (t_start >= t_end) or (t_end > t_vec[-1]):
				tk.messagebox.showerror('Error', 'Time interval is not valid.')
				return False

		if 56 in args:

			try:
				f_min = float(f_min)
				f_max = float(f_max)
			except:
				tk.messagebox.showerror('Error', 'Frecuency interval data are not valid.')
				return False

			if (f_min < 0) or (f_min >= f_max):
				tk.messagebox.showerror('Error', 'Frecuency interval is not valid.')
				return False

		if 7 in args:

			if self.val_opmn2.get() == 'Prony':

				try:
					int(fct_method)
				except:
					tk.messagebox.showerror('Error', 'Modes data are not valid.')
					return False

				fct_method = int(fct_method)
				if fct_method < 0:
					tk.messagebox.showerror('Error', 'Modes data out of range.')
					return False

			if (self.val_opmn2.get() == 'ERA') or (self.val_opmn2.get() == 'Matrix Pencil'):

				try:
					float(fct_method)
				except:
					tk.messagebox.showerror('Error', 'Energy threshold data are not valid.')
					return False

				fct_method = float(fct_method)
				if (fct_method <= 0) or (fct_method > 100.0):
					tk.messagebox.showerror('Error', 'Energy threshold data out of range.')
					return False

		if 9 in args:

			if self.val_W1chb2.get() == True:

				try:
					dwn_spl = float(dwn_spl)					
					if dwn_spl % 1 > 0:
						tk.messagebox.showerror('Error', 'Downsampling factor data are not valid.')
						return False

					dwn_spl = int(dwn_spl)

				except:
					tk.messagebox.showerror('Error', 'Downsampling factor data are not valid.')
					return False

				if dwn_spl < 1:
					tk.messagebox.showerror('Error', 'Downsampling factor out of range.')
					return False

		return True


	def Get_Data(self, file_path):

		Data 	= []
		Y 		= []

		if file_path.endswith('.csv'):
			csvData = csv.reader(open(file_path))
			for column in csvData:
				Data.append(column)

		if file_path.endswith('.xlsx'):
			wb = xlrd.open_workbook(file_path)
			sh = wb.sheet_by_index(0)
			for i in range(sh.nrows):
				Data.append(sh.row_values(i))

		if len(Data) > 0:
			if len(Data[0]) > 1:

				h_vec = Data[0]
				t, y = [], []
				for index, iData in enumerate(Data[1:]):
					t.append(float(iData[0]))
					y.append([])

					for jData in iData[1:]:
						y[index].append(float(jData))

				y_vec = np.array(y)
				t_vec = np.array(t)

				t_vec = t_vec - t_vec[0]

				return Data, h_vec, t_vec, y_vec

		return Data, [], [], []


	def Label_Method(self, *args):
		
		if self.val_opmn2.get() == 'Prony':
			self.lbl4.config(text = 'Modes:')
			self.btn4.config(state = tk.DISABLED)

		if (self.val_opmn2.get() == 'ERA') or  (self.val_opmn2.get() == 'Matrix Pencil'):
			self.lbl4.config(text = 'Energy threshold (%):')
			self.btn4.config(state = tk.NORMAL)

	
	def Downsampling(self, m, t_vec, y_vec):

		t, y = [], []

		for i in range(0 , t_vec.shape[0] , m):
			t.append(t_vec[i])
			y.append(y_vec[i])

		return np.array(t), np.array(y)


	def Normalize(self, signal, n):
	
		signal = signal - np.mean(signal)

		if n == 1:
			for i in range(signal.shape[1]):
				signal[: , i] = signal[: , i] / max(signal[: , i])
			
			return signal

		if n == 2:
			list_max = []
			for i in range(signal.shape[1]):
				list_max.append(max(signal[: , i]))

			return signal / max(list_max)


	def Quit(self, event = None):
		# QUIT
		self.master.destroy()



class Window2():

	# Edit window

	def __init__(self, master, val_opmn3, val_W1ent2,  val_W1ent3,  val_W1ent4, val_W1chb2, val_W1ent9, val_W1chb3, file_path):
		self.master = master
		self.val_opmn3	= val_opmn3
		self.val_W1ent2 = val_W1ent2
		self.val_W1ent3 = val_W1ent3
		self.val_W1ent4 = val_W1ent4
		self.val_W1chb2 = val_W1chb2
		self.val_W1ent9 = val_W1ent9
		self.val_W1chb3 = val_W1chb3
		self.master.title('Data Check')
		self.master.geometry('1600x900')
		self.master.configure(bg = color1)
		self.frame = tk.Frame(self.master)
		self.frame.pack()
		self.Create_Widgets(file_path)


	def Create_Widgets(self, file_path):

		file_name = os.path.basename(file_path)
		Data, h_vec, t_vec, y_vec = self.Get_data(file_path)

		# TITLE LABEL
		self.lbl1 = tk.Label(self.master, text = 'file: ' + file_name, font = (font_text, 14), fg = 'white', bg = color0,  anchor = tk.CENTER)
		self.lbl1.place(height = 25, relwidth = 1)

		# LABELFORM 1 - DATA
		self.lbfrm1 = tk.LabelFrame(self.master, text = 'Data', font = (font_text, 10), fg = 'white', bg = color1)
		self.lbfrm1.place(y = 25, relx = 0.50, relheight = 0.28, relwidth = 0.99, anchor = tk.N)

		self.btn1 = tk.Button(self.lbfrm1, text = 'Select All', font = (font_text, 10), command = lambda: self.click_btn12(True, h_vec, t_vec, y_vec), anchor = tk.CENTER)
		self.btn1.place(y = 0, x = 4, height = 24, relwidth = 0.114, anchor = tk.NW)

		self.btn2 = tk.Button(self.lbfrm1, text = 'Deselect All', font = (font_text, 10), command = lambda: self.click_btn12(False, h_vec, t_vec, y_vec), anchor = tk.CENTER)
		self.btn2.place(y = 28, x = 4, height = 24, relwidth = 0.114, anchor = tk.NW)

		# TREE VIEW FOR SELECTING DATA
		self.trv1 = ttk.Treeview(self.lbfrm1, show = 'headings', selectmode = 'browse', height = 20, columns = ('#1', '#2'))
		self.trv1.place(y = 56, x = 4, relheight = 0.75, relwidth = 0.114, anchor = tk.NW)

		self.trv1.column('#1', minwidth = 0, width = 40, anchor = tk.CENTER, stretch = tk.NO)
		self.trv1.column('#2', minwidth = 0, width = 100, anchor = tk.CENTER)

		self.trv1.heading('#1', text='', anchor=tk.CENTER)
		self.trv1.heading('#2', text='Signal', anchor=tk.CENTER)

		for index, i in enumerate(h_vec[1:]):
			self.trv1.insert('', index, text = True, values = ('✓', i))

		self.vsb = ttk.Scrollbar(self.trv1, orient = 'vertical', command = self.trv1.yview)
		self.vsb.pack(side = 'right', fill = 'y')
		self.trv1.configure(yscrollcommand = self.vsb.set)
		self.trv1.bind('<<TreeviewSelect>>', lambda event: self.Select_trv1(event, h_vec, t_vec, y_vec))

		# TREE VIEW FOR SHOW DATA
		self.trv2 = ttk.Treeview(self.lbfrm1,  show='headings')
		self.trv2.place(relx = 1.00, rely = 0.50, relheight = 1.00, relwidth = .88, anchor = tk.E)
		self.treescrolly = tk.Scrollbar(self.trv2, orient = 'vertical', command = self.trv2.yview)
		self.treescrollx = tk.Scrollbar(self.trv2, orient = 'horizontal', command = self.trv2.xview)
		self.trv2.configure(xscrollcommand = self.treescrollx.set, yscrollcommand = self.treescrolly.set)
		self.treescrolly.pack(side = 'right', fill = tk.Y)
		self.treescrollx.pack(side = 'bottom', fill = tk.X)
		self.Load_trv2(Data)

		# LABELFORM 2 - CHART
		self.lbfrm2 = tk.LabelFrame(self.master, text = 'Chart', font = (font_text, 10), fg = 'white', bg = color1)
		self.lbfrm2.place(relx = 0.50, rely = 0.31, relheight = 0.63, relwidth = 0.99, anchor = tk.N)

		self.lbl3 = tk.Label(self.lbfrm2, text = 'Settings', font = (font_text, 10), fg = 'white', bg = color1, anchor = tk.NW)
		self.lbl3.place(relx = 0.056, rely = 0.040, height = 24, relwidth = 0.11, anchor = tk.CENTER)

		self.lbl4 = tk.Label(self.lbfrm2, text = 'Initial time:', font = (font_text, 10), fg = 'white', bg = color1, anchor = tk.W)
		self.lbl4.place(relx = 0.036, rely = 0.100, height = 24, relwidth = 0.07, anchor = tk.CENTER)

		self.ent1 = tk.Entry(self.lbfrm2, fg = 'white', font = (font_text, 10), bg = color2)
		self.ent1.place(relx = 0.072, rely = 0.100, height = 24, relwidth = 0.038, anchor = tk.W)
		self.ent1.bind('<Return>', lambda event: self.Load_Chart(h_vec, t_vec, y_vec))
		try 	: self.ent1.insert(0, float(self.val_W1ent3.get()))
		except 	: self.ent1.insert(0, float(t_vec[0]))

		self.lbl5 = tk.Label(self.lbfrm2, text = 'Final time:', font = (font_text, 10), fg = 'white', bg = color1, anchor = tk.W)
		self.lbl5.place(relx = 0.036, rely = 0.160, height = 24, relwidth = 0.07, anchor = tk.CENTER)

		self.ent2 = tk.Entry(self.lbfrm2, fg = 'white', font = (font_text, 10), bg = color2)
		self.ent2.place(relx = 0.072, rely = 0.160, height = 24, relwidth = 0.038, anchor = tk.W)
		self.ent2.bind('<Return>', lambda event: self.Load_Chart(h_vec, t_vec, y_vec))
		try 	: self.ent2.insert(0, float(self.val_W1ent4.get()))
		except 	: self.ent2.insert(0, float(t_vec[-1]))

		self.val_W2chb1 = tk.BooleanVar()
		self.val_W2chb1.set(False)
		self.chb1 = tk.Checkbutton(self.lbfrm2, variable = self.val_W2chb1, text = 'Remove mean value', 
			onvalue = True, offvalue = False, fg = 'white', bg = color1, anchor = tk.W, command = lambda: self.Load_Chart(h_vec, t_vec, y_vec))
		self.chb1.config(selectcolor = '#000000')
		self.chb1.place(relx = 0.00, rely = 0.220, height = 24, relwidth = 0.11, anchor = tk.W)

		self.val_W2chb2 = tk.BooleanVar()
		self.val_W2chb2.set(bool(self.val_W1chb3.get()))
		self.chb2 = tk.Checkbutton(self.lbfrm2, variable = self.val_W2chb2, text = 'Normalize all signals', 
			onvalue = True, offvalue = False, fg = 'white', bg = color1, anchor = tk.W, command = lambda: self.Load_Chart(h_vec, t_vec, y_vec))
		self.chb2.config(selectcolor = '#000000')
		self.chb2.place(relx = 0.00, rely = 0.280, height = 24, relwidth = 0.11, anchor = tk.W)

		self.val_W2chb3 = tk.BooleanVar()
		self.val_W2chb3.set(bool(self.val_W1chb2.get()))
		self.chb3 = tk.Checkbutton(self.lbfrm2, variable = self.val_W2chb3, text = 'Downsampling', 
			onvalue = True, offvalue = False, fg = 'white', bg = color1, anchor = tk.W, command = lambda: self.Load_Chart(h_vec, t_vec, y_vec))
		self.chb3.config(selectcolor = '#000000')
		self.chb3.place(relx = 0.00, rely = 0.340, height = 24, relwidth = 0.11, anchor = tk.W)

		self.ent3 = tk.Entry(self.lbfrm2, fg = 'white', font = (font_text, 10), bg = color2)
		self.ent3.place(relx = 0.072, rely = 0.340, height = 24, relwidth = 0.038, anchor = tk.W)
		self.ent3.bind('<Return>', lambda event: self.Load_Chart(h_vec, t_vec, y_vec))
		try 	: self.ent3.insert(0, int(self.val_W1ent9.get()))
		except 	: self.ent3.insert(0, int(1))

		self.lbl6 = tk.Label(self.lbfrm2, font = (font_text, 10), fg = 'white', bg = color1, anchor = tk.W)
		self.lbl6.place(relx = 0.00, rely = 0.40, height = 24, relwidth = 0.11, anchor = tk.W)

		self.Load_Chart(h_vec, t_vec, y_vec)

		# MASTER
		self.btn4 = tk.Button(self.master, text = 'Exit', font = (font_text, 10), command = self.click_btn4, anchor = tk.CENTER)
		self.btn4.place(relx = 0.820, rely = 0.970, height = 26, relwidth = 0.10, anchor = tk.CENTER)

		self.btn5 = tk.Button(self.master, text = 'Accept', font = (font_text, 10), command = lambda: self.click_btn5(h_vec), anchor = tk.CENTER)
		self.btn5.place(relx = 0.930, rely = 0.970, height = 26, relwidth = 0.10, anchor = tk.CENTER)


	def Get_data(self, file_path):

		Data = []
		Y = []

		if file_path.endswith('.csv'):
			csvData = csv.reader(open(file_path))
			for column in csvData:
				Data.append(column)

		if file_path.endswith('.xlsx'):
			wb = xlrd.open_workbook(file_path)
			sh = wb.sheet_by_index(0)
			for i in range(sh.nrows):
				Data.append(sh.row_values(i))

		h_vec = Data[0]
		t, y = [], []
		for index, iData in enumerate(Data[1:]):
			t.append(float(iData[0]))
			y.append([])

			for jData in iData[1:]:
				y[index].append(float(jData))

		y_vec = np.array(y)
		t_vec = np.array(t)

		t_vec = t_vec - t_vec[0]

		return Data, h_vec, t_vec, y_vec

	
	def Select_trv1(self, event, h_vec, t_vec, y_vec):

		dic_item = {}
		for index, item in enumerate(self.trv1.get_children()):
			dic_item[index] = item

		index       = self.trv1.index(self.trv1.selection())
		row_data    = self.trv1.item(dic_item[index])['values']
		row_text    = self.trv1.item(dic_item[index])['text']

		if row_text == True:
			row_data[0] = ''
			text = False
		else:
			row_data[0] = '✓'
			text = True

		self.trv1.delete(dic_item[index])
		self.trv1.insert('', index, text = text, values = row_data)

		self.Load_Chart(h_vec, t_vec, y_vec)


	def Downsampling(self, m, t_vec, y_vec):

		t, y = [], []

		for i in range(0 , t_vec.shape[0] , m):
			t.append(t_vec[i])
			y.append(y_vec[i])

		return np.array(t), np.array(y)


	def Remove_Mean(self, y_vec):

		y = np.zeros([y_vec.shape[0] , y_vec.shape[1]])
		for i in range(y_vec.shape[1]):
			y[: , i] = y_vec[: , i] - np.mean(y_vec[: , i])

		return y


	def Normalize(self, y_vec):

		if y_vec.shape[1] > 0 :

			if self.val_opmn3.get() == 'Single-Channel':

				y_vec = self.Remove_Mean(y_vec)

				for i in range(y_vec.shape[1]):
					y_vec[: , i] = y_vec[: , i] / max(y_vec[: , i])

				return y_vec

			if self.val_opmn3.get() == 'Multi-Channel':

				y_vec = self.Remove_Mean(y_vec)

				max_list = []
				for i in range(y_vec.shape[1]):
					max_list.append(max(y_vec[: , i]))

				y_vec = y_vec / max(max_list)

				return y_vec


	def Selected_hy(self, h_vec, y_vec):

		b = 0
		for item in self.trv1.get_children():
			row_text = self.trv1.item(item)['text']
			if row_text == True:
				b = b + 1

		h = []
		y = np.zeros([y_vec.shape[0] , b])
		b = 0
		for index, item in enumerate(self.trv1.get_children()):
			row_text = self.trv1.item(item)['text']
			if row_text == True:
				h.append(h_vec[1:][index])
				y[: , b] = y_vec[: , index]
				b = b + 1

		return h, y


	def Load_Chart(self, h_vec, t_vec, y_vec):

		h_vec, y_vec = self.Selected_hy(h_vec, y_vec )

		try:
			float(self.ent1.get())
			float(self.ent2.get())

			if (float(self.ent1.get()) < t_vec[0]) or (float(self.ent1.get()) >= float(self.ent2.get())):
				self.ent1.delete(0, 'end')
				self.ent1.insert(0, t_vec[0])

			if (float(self.ent2.get()) <= float(self.ent1.get())) or (float(self.ent2.get()) > t_vec[-1]):
				self.ent2.delete(0, 'end')
				self.ent2.insert(0, t_vec[-1])

		except:
			self.ent1.delete(0, 'end')
			self.ent2.delete(0, 'end')
			self.ent1.insert(0, t_vec[0])
			self.ent2.insert(0, t_vec[-1])

		samples_0 = t_vec.shape[0]
		if self.val_W2chb3.get() == True:
			try:
				int(self.ent3.get())

				if int(self.ent3.get()) < 1:
					self.ent3.delete(0, 'end')
					self.ent3.insert(0, int(1))
			except:
				self.ent3.delete(0, 'end')
				self.ent3.insert(0, int(1))

			m = int(self.ent3.get())
			t_vec, y_vec = self.Downsampling(m, t_vec, y_vec)

		self.lbl6.config(text = 'Samples: ' + str(samples_0) + ' / ' + str(t_vec.shape[0]))

		if self.val_W2chb1.get() == True:
			y_vec = self.Remove_Mean(y_vec)
		
		if self.val_W2chb2.get() == True:
			y_vec = self.Normalize(y_vec)
			self.val_W2chb1.set(True)

		try:
			self.canvas.get_tk_widget().pack_forget()
		except:
			pass

		plt.style.use('dark_background')
		self.fig = plt.Figure(dpi = 100)
		ax = self.fig.add_subplot(111)
		ax.set_xlabel('Time')
		ax.set_ylabel('Magnitude')
		self.fig.subplots_adjust(left = 0.07, bottom = 0.16, right = None, top = 0.95)
		self.canvas = FigureCanvasTkAgg(self.fig, self.lbfrm2)
		self.canvas.draw() 
		self.canvas.get_tk_widget().place(relx = 0.995, rely = 0.50, relheight = 0.98, relwidth = .88, anchor = tk.E)
		self.toolbar = NavigationToolbar2Tk(self.canvas, self.lbfrm2)
		self.toolbar.place(height = 32, relx = 0.995, rely = 0.99, relwidth = .88, anchor = tk.SE)
		self.toolbar.update()

		colors = plt.cm.get_cmap('gist_rainbow', len(h_vec)+1)
		ax.axvline(float(self.ent1.get()), color = 'w', linestyle = 'dashed')
		ax.axvline(float(self.ent2.get()), color = 'w',  linestyle = 'dashed')
		ax.grid(True, color = 'grey',  linestyle = ':', linewidth = 0.75)
		
		try:
			for i in range(y_vec.shape[1]):
				ax.plot(t_vec, y_vec[: , i], label = h_vec[i], lw = '2', color = colors(i))
			ax.legend(loc = 'center right', bbox_to_anchor = (1.12, 0.5))
		except:
			pass
		

	def Load_trv2(self, Data):

		h_vec = Data[0]
		self.trv2['column'] = h_vec

		for i in range(len(h_vec)):
			self.trv2.heading(i, text = h_vec[i])
			self.trv2.column('#' + str(int(i+1)), width = 100, minwidth = 20, stretch = tk.YES, anchor = tk.CENTER)

		for i in range(len(Data)):
			if i > 0: self.trv2.insert('','end', values = Data[i])


	def click_btn12(self, value, h_vec, t_vec, y_vec):
		# DE/SELECT ALL

		for index, item in enumerate(self.trv1.get_children()):

			row_data = self.trv1.item(item)['values']
			row_text = self.trv1.item(item)['text']

			if row_text == (not value):

				if value == True	: row_data[0] = '✓'
				if value == False	: row_data[0] = ''

				self.trv1.delete(item)
				self.trv1.insert('', index, text = value, values = row_data)

		self.Load_Chart(h_vec, t_vec, y_vec)


	def click_btn4(self, event = None):
		# QUIT
		self.master.destroy()


	def click_btn5(self, h_vec):
		# ACCEPT
		res = ''
		for item in self.trv1.get_children():

			row_text = self.trv1.item(item)['text']
			row_data = self.trv1.item(item)['values']

			if row_text == True:
				res = res + ', ' + str(row_data[1])

		self.val_W1ent2.set(res[2:])
		self.val_W1ent3.set(self.ent1.get())
		self.val_W1ent4.set(self.ent2.get())
		self.val_W1chb2.set(self.val_W2chb3.get())
		self.val_W1ent9.set(self.ent3.get())
		self.val_W1chb3.set(self.val_W2chb2.get())

		self.click_btn4()



class Window3():

	def __init__(self, master, val_W1ent7, list_signals, i_channels, t_start, t_end, b_dwnspl, f_dwnspl, energy_th, h_vec, t_vec, y_vec):

		self.master = master
		self.val_W1ent7 = val_W1ent7
		self.master.title('Singular Values')
		self.master.geometry('1200x800')
		self.master.configure(bg = color1)
		self.frame = tk.Frame(self.master)
		self.frame.pack()
		self.Create_Widgets(list_signals, i_channels, t_start, t_end, b_dwnspl, f_dwnspl, h_vec, t_vec, y_vec)



	def Create_Widgets(self, list_signals, i_channels, t_start, t_end, b_dwnspl, f_dwnspl, h_vec, t_vec, y_vec):

		# MASTER
		self.lbl1 = tk.Label(self.master, text = 'Singular Values', font = (font_text, 14), fg = 'white', bg = color0, anchor = tk.CENTER)
		self.lbl1.pack(fill = tk.X)

		# LABERFORM 1 - CHART
		self.lbfrm1 = tk.LabelFrame(self.master, text = 'Chart', font = (font_text, 10), fg = 'white', bg = color1)
		self.lbfrm1.place(y = 30, relx = 0.50, relheight = 0.90, relwidth = 0.99, anchor = tk.N)

		# MASTER
		self.lbl2 = tk.Label(self.master, text = 'Energy threshold (%):', font = (font_text, 10), fg = 'white', bg = color1,  anchor = tk.E)
		self.lbl2.place(relx = 0.280, rely = 0.955, height = 24, relwidth = .20, anchor = tk.NE)

		self.ent1 = tk.Entry(self.master, fg = 'white', font = (font_text, 10), bg = color2)
		self.ent1.place(relx = 0.285, rely = 0.955, height = 24, width = 70)
		self.ent1.insert(0, float(self.val_W1ent7.get()))
		self.ent1.bind('<Return>', lambda event: self.Load_Chart(list_signals, i_channels, svn, r, can))

		# MASTER
		self.btn1 = tk.Button(self.master, text = 'Exit', font = (font_text, 10), command = self.click_btn1, anchor = tk.CENTER)
		self.btn1.place(relx = 0.780, rely = 0.955, height = 26, relwidth = 0.10)

		self.btn2 = tk.Button(self.master, text = 'Accept', font = (font_text, 10), command = self.click_btn2, anchor = tk.CENTER)
		self.btn2.place(relx = 0.890, rely = 0.955, height = 26, relwidth = 0.10)

		N, med, can, r, t_vec, y_vec = self.Signal_prepare(list_signals, t_start, t_end, b_dwnspl, f_dwnspl, h_vec, t_vec, y_vec)
		svn = self.Get_SVN(i_channels, N, med, can, r, t_vec, y_vec)

		self.Load_Chart(list_signals, i_channels, svn, r, can)


	def click_btn1(self):
		# QUIT
		self.master.destroy()


	def click_btn2(self):
		# ACCEPT
		self.val_W1ent7.set(self.ent1.get())
		self.click_btn1()


	def Downsampling(self, m, t_vec, y_vec):

		t, y = [], []

		for i in range(0 , t_vec.shape[0] , m):
			t.append(t_vec[i])
			y.append(y_vec[i])

		return np.array(t), np.array(y)


	def Signal_prepare(self, list_signals, t_start, t_end, b_dwnspl, f_dwnspl, h_vec, t_vec, y_vec):

		for i in range(len(h_vec[1:]), 0, -1):
			if h_vec[i] not in list_signals:
				y_vec = np.delete(y_vec, (i - 1), axis = 1)

		t_start = float(t_start)
		t_end 	= float(t_end)

		dt = float(t_vec[1] - t_vec[0])
		pa = int(round(t_start / dt, 0))
		pb = int(round(t_end / dt, 0))

		t_vec = t_vec[pa : pb]
		y_vec = y_vec[pa : pb , :]

		if b_dwnspl == True:
			t_vec, y_vec = self.Downsampling(f_dwnspl, t_vec, y_vec)

		for i in range(y_vec.shape[1]):
			y_vec[: , i] = y_vec[: , i] - np.mean(y_vec[: , i])

		N 	= len(t_vec)
		med = y_vec.shape[0]
		can = y_vec.shape[1]
		r 	= int(np.around((med / 2.0) - 1.0, 0))

		return N, med, can, r, t_vec, y_vec


	def Get_SVN(self, i_channels, N, med, can, r, t_vec, y_vec):

		if i_channels == 'Single-Channel':

			svn = np.zeros([r, can])

			for c in range(can):

				Y = np.zeros([y_vec.shape[0], 1])
				Y = (y_vec.T)[c]

				Hankel = hankel(Y)
				H0 = Hankel[0:r, 0:r]

				u, s, v = np.linalg.svd(H0)

				svn[: , c] = s

			return svn

		if i_channels == 'Multi-Channel':

			H0 = np.zeros([r * can, r])
			for j in range(r):
				for i in range(r):
					H0[can*i: can*(i+1) , j] = y_vec[j+i+1 , :]

			u, svn, v = np.linalg.svd(H0)

			return svn


	def Load_Chart(self, list_signals, i_channels, svn, r, can):

		if i_channels == 'Single-Channel'	: c_plot = can
		if i_channels == 'Multi-Channel'	: c_plot = 1

		try:
			self.canvas_svn.get_tk_widget().destroy()
			self.toolbar_svn.destroy()
		except:
			pass

		try:
			float(self.ent1.get())

			if float(self.ent1.get()) < 0.0:
				self.ent1.delete(0, 'end')
				self.ent1.insert(0, 0)

			if float(self.ent1.get()) > 100.0:
				self.ent1.delete(0, 'end')
				self.ent1.insert(0, 100.0)

		except:
			self.ent1.delete(0, 'end')
			self.ent1.insert(0, self.val_W1ent7.get())

		enrg_th = float(self.ent1.get())

		x = range(1, svn.shape[0] + 1)

		if enrg_th == 0 	: modes = 0
		if enrg_th == 100.	: modes = r

		colors = plt.cm.get_cmap('gist_rainbow', can+1)

		plt.style.use('dark_background')
		self.fig 	= plt.Figure(dpi = 100)
		ax = self.fig.add_subplot(111)

		for c in range(c_plot):

			if i_channels == 'Single-Channel'	: i_svn = svn[: , c]
			if i_channels == 'Multi-Channel'	: i_svn = svn[:]

			if (enrg_th > 0) and (enrg_th <100):
				sum_s 	= float(sum(i_svn))
				sum_st 	= float(0.0)
				for modes, i in enumerate(i_svn, start = 1):
					sum_st 		= sum_st + i
					pc_sum_st 	= (sum_st / sum_s) * 100.0
					if pc_sum_st >= enrg_th: break

			i_svn = np.log10(i_svn)

			ax.axvline(modes, color = colors(c), linestyle = 'dashed', linewidth = 1.5)
			ax.plot(x, i_svn, color = colors(c), label = list_signals[c] + ' (' + str(int(modes/2.0)) + ')', linewidth = 2.0)

		ax.set_xlabel('Singular Value Number')
		ax.set_ylabel('log10(σ)')
		ax.legend(loc = 'upper right', title = 'Signal (Mode ≤)')
		ax.set_title('SVD from Hankel Matrix')
		ax.grid(True, color = 'grey',  linestyle = ':', linewidth = 0.75)

		self.fig.subplots_adjust(left = 0.08, bottom = 0.16, right = 0.95, top = 0.95)
		self.canvas_svn = FigureCanvasTkAgg(self.fig, self.lbfrm1)
		self.canvas_svn.draw() 
		self.canvas_svn.get_tk_widget().place(y = 5, relx = 0.50, relheight = 0.99, relwidth = 1.00, anchor = tk.N)
		self.toolbar_svn = NavigationToolbar2Tk(self.canvas_svn, self.lbfrm1)
		self.toolbar_svn.place(relx = 0.50, rely = 1.00, relwidth = 1.00, anchor = tk.S)
		self.toolbar_svn.update()


def main():
	root = tk.Tk()
	root.state('zoomed')
	app	= Window1(root)
	root.mainloop()


if __name__ == '__main__':
	main()